import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import range_boundaries
import re
import difflib
from rapidfuzz import fuzz
import xlwings as xw

AdrsDti = r'C:\users\Your Name\Excel File.xlsx'
OpDti = load_workbook(AdrsDti,data_only=True)
BltLnqBc=OpDti['LnqBc']
BltStdKg=OpDti['StdKg']
BltEng=OpDti['Eingabe']
TbLnqBc=BltLnqBc.tables['TbLnqBc']
TbStdKg=BltStdKg.tables["TbStdKg"]
TbEntLst=BltEng.tables["TbEntLst"]

# Die Buchungstexte aus der Tabelle "TbLnqBc" extrahieren
MinSpltLnqBc,MinZilLnqBc,MaxSpltLnqBc,MaxZilLnqBc=range_boundaries(TbLnqBc.ref)
UbschLnqBc=TbLnqBc.column_names
DtnLnqBc=BltLnqBc.iter_rows(min_row=MinZilLnqBc+TbLnqBc.headerRowCount,max_row=MaxZilLnqBc,
                           min_col=MinSpltLnqBc,max_col=MaxSpltLnqBc,values_only=True)
DfPdLnqBc=pd.DataFrame(DtnLnqBc,columns=UbschLnqBc)
LstBcTxt=DfPdLnqBc["Buchungstext"].tolist()

# Die Buchungstexte aus der Tabelle "TbStdKg" extrahieren
MinSpltStdKg,MinZilStdKg,MaxSpltStdKg,MaxZilStdKg=range_boundaries(TbStdKg.ref)
UbschStdKg=TbStdKg.column_names
DtnStdKg=BltStdKg.iter_rows(min_row=MinZilStdKg+TbStdKg.headerRowCount,max_row=MaxZilStdKg,
                                min_col=MinSpltStdKg,max_col=MaxSpltStdKg,values_only=True)
DfPdStdKg=pd.DataFrame(DtnStdKg,columns=UbschStdKg)
LstStdKgTxt=DfPdStdKg["Buchungstext"].tolist()
LstStdKgKg=DfPdStdKg["Kg"].tolist()

# Die Entlist aus der Tabelle "TbEntLst" extrahieren
MinSpltEntLst,MinZilEntLst,MaxSpltEntLst,MaxZilEntLst=range_boundaries(TbEntLst.ref)
UbschEntLst=TbEntLst.column_names
DtnEntLst=BltEng.iter_rows(min_row=MinZilEntLst+TbEntLst.headerRowCount,max_row=MaxZilEntLst,
                                min_col=MinSpltEntLst,max_col=MaxSpltEntLst,values_only=True)
DfPdEntLst=pd.DataFrame(DtnEntLst,columns=UbschEntLst)
LstEntLst=DfPdEntLst["Entlist"].tolist()

# Die extrahierten Listen anzuzeigen
print(LstBcTxt)
print(LstStdKgTxt)
print(LstStdKgKg)
print(LstEntLst)

# Die benötigten Austellungen vorzubereiten
LstBcTxtSb=[]
LstBcTxtSrt=[]
LstDflbTxt=[]
LstFuzzTxt=[]
LstMxDflb=[]
LstMxFuzz=[]
LstDflbIdx=[]
LstFuzzIdx=[]
LstDflbKg=[]
LstFuzzKg=[]

# Die Entlist aus LstEntLst zu entfernen und die Texte aus LstBcTxt zur LstStdKgTxt zu vergleichen
for i,bt in enumerate(LstBcTxt):
    for Ent in LstEntLst:
        bt=re.sub(r"\b" + re.escape(Ent) + r"\b", " ", bt, flags=re.IGNORECASE)
    bt=re.sub(r"[^A-Za-zÄÖÜäöüß\s]", "",bt).strip()
    LstBcTxtSb.append(bt)
    bt=" ".join(sorted(bt.split(), key=str.lower))
    LstBcTxtSrt.append(bt)
    LstBcTxt[i]=bt
    LstDflbStz=[]
    LstFuzzStz=[]
    for sbt in LstStdKgTxt:
        if sbt is None or bt==" ":
            LstDflbStz.append(0)
            LstFuzzStz.append(0)
        else:
            LstDflbStz.append(difflib.SequenceMatcher(None, bt, sbt).ratio())
            LstFuzzStz.append((fuzz.token_sort_ratio(bt, sbt)))
    LstDflbIdx.append(LstDflbStz.index(max(LstDflbStz))+1)
    LstFuzzIdx.append(LstFuzzStz.index(max(LstFuzzStz))+1)
    LstMxDflb.append(max(LstDflbStz))
    LstMxFuzz.append(max(LstFuzzStz))
    LstDflbTxt.append(LstStdKgTxt[LstDflbIdx[i]-1])
    LstFuzzTxt.append(LstStdKgTxt[LstFuzzIdx[i]-1])
    LstDflbKg.append(LstStdKgKg[LstDflbIdx[i]-1])
    LstFuzzKg.append(LstStdKgKg[LstFuzzIdx[i]-1])
    print(i, bt)
    print(LstDflbStz)
    print(LstFuzzStz)
print(LstDflbIdx)
print(LstFuzzIdx)
print(LstMxDflb)
print(LstMxFuzz)
print(LstDflbTxt)
print(LstFuzzTxt)
print(LstDflbKg)
print(LstFuzzKg)
print(LstBcTxtSb)
print(LstBcTxtSrt)

# Die Ergebnisse in die Tabelle "TbLnqBc" zu schreiben
xlwDti=xw.Book(AdrsDti)
xlwBlt=xlwDti.sheets["LnqBc"]
xlwTb=xlwBlt.tables["TbLnqBc"]

SpltNrIdxDf=xlwTb.header_row_range.value.index("IdxDf")
SpltNrIdxFz=xlwTb.header_row_range.value.index("IdxFz")
SpltNrStDf=xlwTb.header_row_range.value.index("StDf")
SpltNrStFz=xlwTb.header_row_range.value.index("StFz")
SpltNrRtDf=xlwTb.header_row_range.value.index("RtDf")
SpltNrRtFz=xlwTb.header_row_range.value.index("RtFz")
SpltNrStDfBt=xlwTb.header_row_range.value.index("StDfBt")
SpltNrStFzBt=xlwTb.header_row_range.value.index("StFzBt")
SpltNrBt=xlwTb.header_row_range.value.index("Bt")
SpltNrSrtdBt=xlwTb.header_row_range.value.index("SrtdBt")

print(SpltNrIdxDf)
print(SpltNrIdxFz)
print(SpltNrStDf)
print(SpltNrStFz)
print(SpltNrRtDf)
print(SpltNrRtFz)
print(SpltNrStDfBt)
print(SpltNrStFzBt)
print(SpltNrBt)
print(SpltNrSrtdBt)

xlwTb.data_body_range.columns[SpltNrIdxDf].options(transpose=True).value = LstDflbIdx
xlwTb.data_body_range.columns[SpltNrIdxFz].options(transpose=True).value = LstFuzzIdx
xlwTb.data_body_range.columns[SpltNrStDf].options(transpose=True).value = LstDflbKg
xlwTb.data_body_range.columns[SpltNrStFz].options(transpose=True).value = LstFuzzKg
xlwTb.data_body_range.columns[SpltNrRtDf].options(transpose=True).value = LstMxDflb
xlwTb.data_body_range.columns[SpltNrRtFz].options(transpose=True).value = LstMxFuzz
xlwTb.data_body_range.columns[SpltNrStDfBt].options(transpose=True).value = LstDflbTxt
xlwTb.data_body_range.columns[SpltNrStFzBt].options(transpose=True).value = LstFuzzTxt
xlwTb.data_body_range.columns[SpltNrBt].options(transpose=True).value = LstBcTxtSb
xlwTb.data_body_range.columns[SpltNrSrtdBt].options(transpose=True).value = LstBcTxtSrt

# Das Ergebnis anzuzeigen
print(PdDti)
print(TbLnqBc.ref)
print(MinZilLnq,MaxZilLnq, MinSpltLnq, MaxSpltLnq)
print(UbschLnqBc)
print(TbEntLst)
print(SpltBcTxt)
print(UbschStdKg)
print(len(LstDflbTxt))
